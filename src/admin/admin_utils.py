# Стандартные библиотеки Python
import os
import shutil
from datetime import datetime

# Библиотеки для работы с данными и Excel
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Border,
    Side
)
from openpyxl.utils import get_column_letter

# Библиотеки Aiogram
from aiogram import types
from aiogram.fsm.context import FSMContext
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton

# Локальные импорты
from src.messages import BUTTON_BACK
from src.database import get_user, get_all_reviews, get_all_questions
from .admin_keyboards import get_admin_menu_keyboard
from .admin_messages import ADMIN_MENU_TEXT

async def show_admin_menu(message, user_id: int, is_bot: bool):
    """
    Показывает меню администратора.
    
    Args:
        message: Объект сообщения или callback
        user_id: ID пользователя
        is_bot: Флаг, указывающий, что сообщение отправлено ботом
    """
    # Получаем уровень доступа администратора из базы данных
    user = await get_user(user_id)
    admin_level = user[2] if user else 0  # user[2] - это поле admin_rights в базе данных
    
    if is_bot:
        await message.edit_text(
            ADMIN_MENU_TEXT,
            reply_markup=get_admin_menu_keyboard(admin_level)
        )
    else:
        await message.answer(
            ADMIN_MENU_TEXT,
            reply_markup=get_admin_menu_keyboard(admin_level)
        )



async def export_reviews_excel(message: types.Message | types.CallbackQuery):
    """
    Экспорт отзывов в Excel.
    """
    if isinstance(message, types.CallbackQuery):
        await message.message.edit_text("⏳ Подготовка экспорта отзывов...\n\nПожалуйста подождите, это может занять некоторое время...")
    else:
        await message.edit_text("⏳ Подготовка экспорта отзывов...\n\nПожалуйста подождите, это может занять некоторое время...")

    # Получаем все отзывы из базы данных
    reviews = await get_all_reviews()
    
    if not reviews:
        if isinstance(message, types.CallbackQuery):
            await message.message.edit_text(
                "❌ В базе данных пока нет отзывов",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=BUTTON_BACK, callback_data="back_to_main")]])
            )
        else:
            await message.edit_text(
                "❌ В базе данных пока нет отзывов",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=BUTTON_BACK, callback_data="back_to_main")]])
            )
        return

    # Создаем директорию для экспорта, если её нет
    export_dir = 'exports'
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    
    # Генерируем имя файла с текущей датой
    filename = f'отзывы_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)
    
    # Копируем шаблон
    template_path = os.path.join('templates', 'reviews_template.xlsx')
    shutil.copy2(template_path, filepath)
    
    # Загружаем книгу Excel
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Создаем стиль границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Счетчик отзывов без ответа
    unanswered_count = 0
    
    # Заполняем данные
    for i, review in enumerate(reviews, start=2):  # Начинаем с 2-й строки (после заголовка)
        ws.cell(row=i, column=1, value=review[0])  # ID
        ws.cell(row=i, column=2, value=review[1])  # User ID
        ws.cell(row=i, column=3, value=review[2])  # Username
        ws.cell(row=i, column=4, value=review[3])  # Rating
        ws.cell(row=i, column=5, value=review[4])  # Review Text
        ws.cell(row=i, column=6, value=review[5])  # Admin Response
        # Форматируем дату
        created_at = datetime.strptime(str(review[6]), "%Y-%m-%d %H:%M:%S.%f")
        ws.cell(row=i, column=7, value=created_at.strftime("%d.%m.%Y %H:%M"))  # Created At
        
        # Применяем границы ко всем ячейкам в строке
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = thin_border
            
        # Считаем отзывы без ответа
        if not review[5]:  # Если нет ответа администратора
            unanswered_count += 1
    
    # Добавляем статистику
    ws['H2'] = f"{len(reviews)}"
    ws['I2'] = f"{unanswered_count}"
    
    # Добавляем фильтры
    ws.auto_filter.ref = f"A1:G{len(reviews) + 1}"
    
    # Сохраняем изменения
    wb.save(filepath)
    
    # Отправляем файл
    if isinstance(message, types.CallbackQuery):
        await message.message.edit_text("✅ Экспорт отзывов успешно завершен", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="↩️ Вернуться в меню", callback_data="back_to_main")]
        ]))
        await message.message.answer_document(
            types.FSInputFile(filepath)
        )
    else:
        await message.edit_text("✅ Экспорт отзывов успешно завершен", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="↩️ Вернуться в меню", callback_data="back_to_main")]
        ]))
        await message.answer_document(
            types.FSInputFile(filepath)
        )

async def export_questions_excel(message: types.Message | types.CallbackQuery):
    """
    Экспорт вопросов в Excel.
    """
    if isinstance(message, types.CallbackQuery):
        await message.message.edit_text("⏳ Подготовка экспорта вопросов...\n\nПожалуйста подождите, это может занять некоторое время...")
    else:
        await message.edit_text("⏳ Подготовка экспорта вопросов...\n\nПожалуйста подождите, это может занять некоторое время...")

    # Получаем все вопросы из базы данных
    questions = await get_all_questions()
    
    if not questions:
        if isinstance(message, types.CallbackQuery):
            await message.message.edit_text(
                "❌ В базе данных пока нет вопросов",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=BUTTON_BACK, callback_data="back_to_main")]])
            )
        else:
            await message.edit_text(
                "❌ В базе данных пока нет вопросов",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=BUTTON_BACK, callback_data="back_to_main")]])
            )
        return

    # Создаем директорию для экспорта, если её нет
    export_dir = 'exports'
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    
    # Генерируем имя файла с текущей датой
    filename = f'вопросы_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)
    
    # Копируем шаблон
    template_path = os.path.join('templates', 'questions_template.xlsx')
    shutil.copy2(template_path, filepath)
    
    # Загружаем книгу Excel
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Создаем стиль границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Счетчик вопросов без ответа
    unanswered_count = 0
    
    # Заполняем данные
    for i, question in enumerate(questions, start=2):  # Начинаем с 2-й строки (после заголовка)
        ws.cell(row=i, column=1, value=question[0])  # ID
        ws.cell(row=i, column=2, value=question[1])  # User ID
        ws.cell(row=i, column=3, value=question[2])  # Username
        ws.cell(row=i, column=4, value=question[3])  # Question Text
        ws.cell(row=i, column=5, value=question[4])  # Admin Response
        # Форматируем дату
        created_at = datetime.strptime(str(question[5]), "%Y-%m-%d %H:%M:%S.%f")
        ws.cell(row=i, column=6, value=created_at.strftime("%d.%m.%Y %H:%M"))  # Created At
        
        # Применяем границы ко всем ячейкам в строке
        for col in range(1, 7):
            ws.cell(row=i, column=col).border = thin_border
            
        # Считаем вопросы без ответа
        if not question[4]:  # Если нет ответа администратора
            unanswered_count += 1
    
    # Добавляем статистику
    ws['H2'] = f"{len(questions)}"
    ws['I2'] = f"{unanswered_count}"
    
    # Добавляем фильтры
    ws.auto_filter.ref = f"A1:F{len(questions) + 1}"
    
    # Сохраняем изменения
    wb.save(filepath)
    
    # Отправляем файл
    if isinstance(message, types.CallbackQuery):
        await message.message.edit_text("✅ Экспорт вопросов успешно завершен", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="↩️ Вернуться в меню", callback_data="back_to_main")]
        ]))
        await message.message.answer_document(
            types.FSInputFile(filepath)
        )
    else:
        await message.edit_text("✅ Экспорт вопросов успешно завершен", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="↩️ Вернуться в меню", callback_data="back_to_main")]
        ]))
        await message.answer_document(
            types.FSInputFile(filepath)
        )

